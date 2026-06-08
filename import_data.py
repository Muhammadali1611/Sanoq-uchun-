"""
Klent va tovar bazasini Excel fayllardan botning bazasiga yuklaydi.
Ishlatish:  python import_data.py
Fayllar:    data/klentlar.xlsx , data/tovarlar.xlsx
"""
import asyncio
from pathlib import Path

from openpyxl import load_workbook

import database as db

DATA = Path(__file__).resolve().parent / "data"
PRODUCT_UNIT = "dona"   # barcha tovarlar uchun standart birlik


def parse_clients(path):
    """Sariq rangli qator = region sarlavhasi; qolganlari shu regiondagi klent."""
    wb = load_workbook(path)
    ws = wb.worksheets[0]
    region = None
    out = []
    for r in range(1, ws.max_row + 1):
        cell = ws.cell(r, 1)
        val = cell.value
        if val is None or not str(val).strip():
            continue
        name = str(val).strip()
        is_region = cell.fill is not None and cell.fill.patternType is not None
        if is_region:
            region = name
        else:
            out.append((name, region))
    return out


def parse_products(path):
    """Лист2 dagi 'Товар / Цена' ro'yxati."""
    wb = load_workbook(path, data_only=True)
    ws = wb["Лист2"] if "Лист2" in wb.sheetnames else wb.worksheets[0]
    out = []
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, 1).value
        price = ws.cell(r, 2).value
        if name and str(name).strip() and price is not None:
            out.append((str(name).strip(), float(price)))
    return out


async def main():
    await db.init_db()

    clients = parse_clients(DATA / "klentlar.xlsx")
    products = parse_products(DATA / "tovarlar.xlsx")

    # Takror yuklanmasligi uchun mavjudlarini tekshiramiz
    existing_clients = {c["name"] for c in await db.get_clients()}
    existing_products = {p["name"] for p in await db.get_products()}

    added_c = 0
    for name, region in clients:
        if name not in existing_clients:
            await db.add_client(name, region=region)
            added_c += 1

    added_p = 0
    for name, price in products:
        if name not in existing_products:
            await db.add_product(name, PRODUCT_UNIT, price)
            added_p += 1

    regions = await db.get_regions()
    print(f"✅ Klentlar: {added_c} ta qo'shildi ({len(regions)} region)")
    print(f"✅ Tovarlar: {added_p} ta qo'shildi")
    print("\nRegionlar:")
    for rg in regions:
        print(f"  • {rg['region']}: {rg['cnt']} ta")


if __name__ == "__main__":
    asyncio.run(main())
