"""
Excel hisobotlar (openpyxl).
Kam qolgan yuk (qoldiq <= LOW_STOCK_QTY) → qizil rangda.
"""
import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import LOW_STOCK_QTY


# --- Umumiy stillar ---
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F4E78")
BOLD = Font(bold=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center")
MONEY_FMT = "#,##0"
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _style_header(ws, row_idx, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def _autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _save(wb) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ---------------------------------------------------------------------------
# KUNLIK HISOBOT
# ---------------------------------------------------------------------------
def build_daily_report(date_str: str, rows: list) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Kunlik sanash"

    ws.merge_cells("A1:H1")
    ws["A1"] = f"📋 KUNLIK SANASH HISOBOTI  —  {date_str}"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = LEFT

    headers = ["Agent", "Klent", "Tovar", "Birlik", "Qoldiq", "Narx (so'm)", "Summa (so'm)", "Vaqt"]
    hr = 3
    for c, h in enumerate(headers, start=1):
        ws.cell(row=hr, column=c, value=h)
    _style_header(ws, hr, len(headers))

    r = hr + 1
    total = 0.0
    for row in rows:
        summa = row["quantity"] * row["price"]
        total += summa
        vals = [
            row["agent_name"], row["client_name"], row["product_name"],
            row["unit"], row["quantity"], row["price"], summa,
            row["created_at"][11:16],
        ]
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = BORDER
            if c in (6, 7):
                cell.number_format = MONEY_FMT
            if c == 5:
                cell.alignment = CENTER
        # Kam qolgan → qizil
        if row["quantity"] <= LOW_STOCK_QTY:
            for c in range(1, 9):
                ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor="FFC7CE")
        r += 1

    # Jami qatori
    ws.cell(row=r, column=6, value="JAMI:").font = BOLD
    tcell = ws.cell(row=r, column=7, value=total)
    tcell.font = BOLD
    tcell.number_format = MONEY_FMT
    tcell.fill = PatternFill("solid", fgColor="FFF2CC")

    _autosize(ws, [18, 22, 26, 8, 10, 14, 16, 8])
    ws.freeze_panes = "A4"
    if rows:
        ws.auto_filter.ref = f"A{hr}:H{r-1}"
    else:
        ws.cell(row=4, column=1, value="Bu sanada sanash bo'lmagan.")
    return _save(wb)


# ---------------------------------------------------------------------------
# OYLIK HISOBOT (3 varaq) — qoldiqlar va kam qolganlar qizil
# ---------------------------------------------------------------------------
from config import LOW_STOCK_QTY

RED_FILL = PatternFill("solid", fgColor="FFC7CE")      # kam qolgan
LIGHT_RED = PatternFill("solid", fgColor="FFCCCC")


def _latest_snapshot(period_rows: list):
    """Har bir klent x tovar uchun oxirgi sanash qoldig'ini oladi."""
    best = {}  # (client_name, product_name) -> row
    for r in period_rows:
        key = (r["client_name"], r["product_name"])
        prev = best.get(key)
        if prev is None or (r["count_date"], r["created_at"]) > (prev["count_date"], prev["created_at"]):
            best[key] = r
    return list(best.values())


def build_monthly_report(title: str, period_rows: list) -> bytes:
    wb = Workbook()
    snapshot = _latest_snapshot(period_rows)

    # ===== Varaq 1: Klent qoldiqlari =====
    ws = wb.active
    ws.title = "Klent qoldiqlari"
    ws.merge_cells("A1:E1")
    ws["A1"] = f"📦 KLENT QOLDIQLARI  —  {title}"
    ws["A1"].font = TITLE_FONT

    headers = ["Klent", "Oxirgi sana", "Tovarlar soni", "Jami qoldiq (dona)", "Jami summa (so'm)"]
    hr = 3
    for c, h in enumerate(headers, start=1):
        ws.cell(row=hr, column=c, value=h)
    _style_header(ws, hr, len(headers))

    # Klent bo'yicha guruhlash
    from collections import defaultdict
    clients = defaultdict(lambda: {"date": "", "products": 0, "qty": 0.0, "value": 0.0})
    for r in snapshot:
        cn = r["client_name"]
        clients[cn]["qty"] += r["quantity"]
        clients[cn]["value"] += r["quantity"] * r["price"]
        clients[cn]["products"] += 1
        if r["count_date"] > clients[cn]["date"]:
            clients[cn]["date"] = r["count_date"]

    row = hr + 1
    for cn in sorted(clients.keys()):
        c_data = clients[cn]
        vals = [cn, c_data["date"], c_data["products"],
                round(c_data["qty"], 1), round(c_data["value"])]
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.border = BORDER
            if c == 5:
                cell.number_format = MONEY_FMT
            if c in (3, 4):
                cell.alignment = CENTER
        row += 1

    _autosize(ws, [28, 13, 14, 16, 18])
    ws.freeze_panes = "A4"
    if clients:
        ws.auto_filter.ref = f"A{hr}:E{row-1}"

    # ===== Varaq 2: Tovar bo'yicha (qizil = kam qolgan) =====
    ws2 = wb.create_sheet("Tovar bo'yicha")
    ws2.merge_cells("A1:E1")
    ws2["A1"] = f"📦 TOVAR BO'YICHA  —  {title}"
    ws2["A1"].font = TITLE_FONT
    h2 = ["Klent", "Tovar", "Qoldiq", "Narx (so'm)", "Summa (so'm)"]
    for c, h in enumerate(h2, start=1):
        ws2.cell(row=3, column=c, value=h)
    _style_header(ws2, 3, len(h2))

    rr = 4
    # Klent bo'yicha tartiblash
    sorted_snap = sorted(snapshot, key=lambda x: (x["client_name"], x["product_name"]))
    for r in sorted_snap:
        summa = r["quantity"] * r["price"]
        vals = [r["client_name"], r["product_name"], r["quantity"], r["price"], summa]
        for c, v in enumerate(vals, start=1):
            cell = ws2.cell(row=rr, column=c, value=v)
            cell.border = BORDER
            if c in (4, 5):
                cell.number_format = MONEY_FMT
            if c == 3:
                cell.alignment = CENTER
        # Kam qolgan → qizil
        if r["quantity"] <= LOW_STOCK_QTY:
            for c in range(1, 6):
                ws2.cell(row=rr, column=c).fill = RED_FILL
        rr += 1

    # Izoh
    rr += 1
    ws2.cell(row=rr, column=1, value="Izoh:").font = BOLD
    rr += 1
    iz = ws2.cell(row=rr, column=1, value=f"🔴 Qizil = qoldiq {LOW_STOCK_QTY} dona va undan kam")
    iz.fill = RED_FILL

    _autosize(ws2, [28, 28, 10, 14, 18])
    ws2.freeze_panes = "A4"
    if sorted_snap:
        ws2.auto_filter.ref = f"A3:E{rr-2}"

    # ===== Varaq 3: Batafsil sanashlar =====
    ws3 = wb.create_sheet("Batafsil sanashlar")
    ws3.merge_cells("A1:G1")
    ws3["A1"] = f"🗂 BATAFSIL SANASHLAR  —  {title}"
    ws3["A1"].font = TITLE_FONT
    h3 = ["Sana", "Agent", "Klent", "Tovar", "Qoldiq", "Narx (so'm)", "Summa (so'm)"]
    for c, h in enumerate(h3, start=1):
        ws3.cell(row=3, column=c, value=h)
    _style_header(ws3, 3, len(h3))
    rr = 4
    for row in period_rows:
        summa = row["quantity"] * row["price"]
        vals = [row["count_date"], row["agent_name"], row["client_name"],
                row["product_name"], row["quantity"], row["price"], summa]
        for c, v in enumerate(vals, start=1):
            cell = ws3.cell(row=rr, column=c, value=v)
            cell.border = BORDER
            if c in (6, 7):
                cell.number_format = MONEY_FMT
            if c == 5:
                cell.alignment = CENTER
        # Kam qolgan qizil
        if row["quantity"] <= LOW_STOCK_QTY:
            ws3.cell(row=rr, column=5).fill = RED_FILL
        rr += 1
    _autosize(ws3, [12, 18, 28, 28, 10, 14, 16])
    ws3.freeze_panes = "A4"
    if period_rows:
        ws3.auto_filter.ref = f"A3:G{rr-1}"

    return _save(wb)
